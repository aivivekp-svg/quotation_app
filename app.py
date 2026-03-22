"""
V. Purohit & Associates — Quotation Generator
Phase 2: Google Sheets | Client Master | Saved Quotations | Fee Master | Reports
"""

import io
import re
import uuid
from datetime import datetime, timedelta
from typing import Optional

import pandas as pd
import streamlit as st
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    BaseDocTemplate, Frame, PageTemplate,
    Paragraph, Spacer, Table, TableStyle, PageBreak
)

# ── Constants ─────────────────────────────────────────────────────────────────

APP_TITLE      = "Quotation Generator – V. Purohit & Associates"
BRAND_BLUE_HEX = "#0F4C81"
BRAND_LIGHT    = "#E8F0FB"
GST_RATE       = 18

ACCOUNTING_PLANS = [
    "Monthly Accounting", "Quarterly Accounting",
    "Half Yearly Accounting", "Annual Accounting",
]
EVENT_SERVICE    = "EVENT BASED FILING"
PT_SERVICE       = "PROFESSION TAX RETURNS"
FORCE_EVENT_SUBS = {
    "FILING OF TDS RETURN IN FORM 26QB",
    "FILING OF TDS RETURN IN FORM 26QC",
    "FILING OF TDS RETURN IN FORM 27Q",
}
ALIAS_DUPLICATE  = {"LIMITED COMPANY": "PRIVATE LIMITED"}
ACRONYMS = ["GST","GSTR","PTEC","PTRC","ADT","ROC","TDS",
            "AOC","MGT","26QB","26QC","DIR","MSME","KYC"]
SUBSERVICE_RENAMES = {
    "CHANGE OF ADDRESS IN GST": "GST Amendment",
    "DIR 12": "DIR 12",
    "MSME APPLICATION": "MSME Application",
    "ROC E-KYC FOR DIRECTORS": "ROC E-KYC For Directors",
}

# Google Sheets tab names — must match exactly in your Sheet
SH_CLIENT  = "ClientMaster"
SH_QUOTES  = "SavedQuotations"
SH_FEES    = "FeeMaster"
SH_REPORTS = "QuoteReports"


# ── Google Sheets helpers ─────────────────────────────────────────────────────

@st.cache_resource
def get_gc():
    import gspread
    from google.oauth2.service_account import Credentials
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    info  = dict(st.secrets["gcp_service_account"])
    creds = Credentials.from_service_account_info(info, scopes=scopes)
    return gspread.authorize(creds)


def open_ws(tab: str):
    gc = get_gc()
    sh = gc.open(st.secrets["spreadsheet_name"])
    return sh.worksheet(tab)


def sheet_to_df(tab: str) -> pd.DataFrame:
    ws   = open_ws(tab)
    data = ws.get_all_records()
    return pd.DataFrame(data) if data else pd.DataFrame()


def df_to_sheet(tab: str, df: pd.DataFrame):
    ws = open_ws(tab)
    ws.clear()
    ws.update([df.columns.tolist()] +
              df.fillna("").astype(str).values.tolist())


def append_row_ws(tab: str, row: list):
    open_ws(tab).append_row(
        [str(x) for x in row], value_input_option="USER_ENTERED")


# ── Client Master ─────────────────────────────────────────────────────────────

CLIENT_COLS = ["ClientID","ClientName","ClientType",
               "Address","Email","Phone","CreatedOn"]

def load_clients() -> pd.DataFrame:
    try:
        df = sheet_to_df(SH_CLIENT)
        return df if not df.empty else pd.DataFrame(columns=CLIENT_COLS)
    except Exception:
        return pd.DataFrame(columns=CLIENT_COLS)


def save_new_client(rec: dict):
    append_row_ws(SH_CLIENT, [rec.get(c,"") for c in CLIENT_COLS])


def overwrite_clients(df: pd.DataFrame):
    df_to_sheet(SH_CLIENT, df[CLIENT_COLS])


# ── Fee Master ────────────────────────────────────────────────────────────────

FEE_COLS = ["Service","SubService","ClientType","Applicable","FeeINR"]

def load_fee_master_raw() -> pd.DataFrame:
    try:
        df = sheet_to_df(SH_FEES)
        if df.empty:
            return pd.DataFrame(columns=FEE_COLS)
        df["Applicable"] = df["Applicable"].astype(str).str.upper().isin(["TRUE","1","YES"])
        df["FeeINR"]     = pd.to_numeric(df["FeeINR"], errors="coerce").fillna(0.0)
        return df
    except Exception:
        return pd.DataFrame(columns=FEE_COLS)


def save_fee_master(df: pd.DataFrame):
    out = df.copy()
    out["Applicable"] = out["Applicable"].map(lambda x: "TRUE" if x else "FALSE")
    df_to_sheet(SH_FEES, out[FEE_COLS])


@st.cache_data(ttl=300)
def cached_fee_master():
    df = load_fee_master_raw()
    for new_t, base_t in ALIAS_DUPLICATE.items():
        base = normalize_str(base_t)
        new  = normalize_str(new_t)
        rows = df[df["ClientType"].map(normalize_str) == base].copy()
        if not rows.empty:
            rows["ClientType"] = new
            df = pd.concat([df, rows], ignore_index=True)
    df["Service"]    = df["Service"].map(normalize_str)
    df["SubService"] = df["SubService"].map(
        lambda v: normalize_str(v) if pd.notna(v) else "")
    df["ClientType"] = df["ClientType"].map(normalize_str)
    return df


# ── Saved Quotations ──────────────────────────────────────────────────────────

QUOTE_COLS = [
    "QuoteNo","Date","ClientID","ClientName","ClientType",
    "Service","Details","FeeINR","Discount_pct","DiscountReason",
    "Subtotal","GSTAmt","GrandTotal","ProposalStart","Status","FY",
]
REPORT_COLS = [
    "QuoteNo","Date","ClientName","ClientType",
    "Subtotal","Discount_pct","GSTAmt","GrandTotal","Status","FY",
]

def save_quotation(quote_no, client_id, client_name, client_type,
                   df_main, subtotal, discount_pct, discount_reason,
                   gst_amt, grand, proposal_start, status="Sent"):
    today = datetime.now().strftime("%d-%b-%Y")
    fy    = get_fy(datetime.now())
    ws    = open_ws(SH_QUOTES)
    for _, r in df_main.iterrows():
        ws.append_row([
            quote_no, today, client_id, client_name, client_type,
            r.get("Service",""), r.get("Details",""),
            parse_inr(r.get("Annual Fees (Rs.)","0")),
            discount_pct, discount_reason,
            subtotal, gst_amt, grand, proposal_start, status, fy,
        ], value_input_option="USER_ENTERED")
    try:
        append_row_ws(SH_REPORTS, [
            quote_no, today, client_name, client_type,
            subtotal, discount_pct, gst_amt, grand, status, fy,
        ])
    except Exception:
        pass


def load_quotations() -> pd.DataFrame:
    try:
        df = sheet_to_df(SH_QUOTES)
        if df.empty:
            return pd.DataFrame(columns=QUOTE_COLS)
        for col in ["FeeINR","GrandTotal","Subtotal","GSTAmt"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        return df
    except Exception:
        return pd.DataFrame(columns=QUOTE_COLS)


def load_reports() -> pd.DataFrame:
    try:
        df = sheet_to_df(SH_REPORTS)
        if df.empty:
            return pd.DataFrame(columns=REPORT_COLS)
        for col in ["Subtotal","GSTAmt","GrandTotal"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        return df
    except Exception:
        return pd.DataFrame(columns=REPORT_COLS)


# ── Utility functions ─────────────────────────────────────────────────────────

def normalize_str(x):
    return (x or "").strip().upper()


def title_with_acronyms(text: str) -> str:
    if not text:
        return ""
    t = " ".join(str(text).split()).title()
    t = re.sub(r"\bOf\b", "of", t, flags=re.IGNORECASE)
    for token in ACRONYMS:
        t = re.sub(rf"\b{re.escape(token)}\b", token, t, flags=re.IGNORECASE)
    return t


def svc_display(raw, pretty):
    if raw == "FILING OF GSTR RETURNS":
        return "Filing of GST Returns"
    return pretty


def sub_display(raw, pretty):
    return SUBSERVICE_RENAMES.get(raw, pretty)


def money_inr(n) -> str:
    try:
        n = float(n)
    except Exception:
        return "0"
    neg = n < 0
    n   = abs(int(round(n)))
    s   = str(n)
    if len(s) <= 3:
        res = s
    else:
        res = s[-3:]
        s   = s[:-3]
        while len(s) > 2:
            res = s[-2:] + "," + res
            s   = s[:-2]
        if s:
            res = s + "," + res
    return ("-" + res) if neg else res


def parse_inr(s) -> float:
    if s is None:
        return 0.0
    if isinstance(s, (int, float)):
        return float(s)
    try:
        return float(str(s).strip().replace(",",""))
    except Exception:
        return 0.0


def validity_date(days=15) -> str:
    return (datetime.now() + timedelta(days=days)).strftime("%d-%b-%Y")


def get_fy(dt: datetime) -> str:
    y = dt.year
    return f"FY {y-1}-{str(y)[2:]}" if dt.month < 4 else f"FY {y}-{str(y+1)[2:]}"


def generate_quote_no() -> str:
    """Generate sequential quote number: VPA/25-26/001"""
    fy     = get_fy(datetime.now())            # e.g. "FY 2025-26"
    fy_short = fy.replace("FY ","").replace("20","")  # e.g. "25-26"
    try:
        rpt = sheet_to_df(SH_REPORTS)
        if rpt.empty:
            count = 1
        else:
            fy_rows = rpt[rpt["FY"] == fy] if "FY" in rpt.columns else rpt
            count   = len(fy_rows) + 1
    except Exception:
        count = 1
    return f"VPA/{fy_short}/{count:03d}"


def validate_email(e: str) -> bool:
    if not e.strip():
        return True
    return bool(re.match(r"^[\w\.-]+@[\w\.-]+\.\w{2,}$", e.strip()))


def validate_phone(p: str) -> bool:
    if not p.strip():
        return True
    d = re.sub(r"[\s\-\+\(\)]", "", p)
    return d.isdigit() and 7 <= len(d) <= 15


def compute_totals(df: pd.DataFrame, discount_pct: float):
    subtotal    = float(df["Annual Fees (Rs.)"].apply(parse_inr).sum())
    disc_amt    = round(subtotal * (discount_pct or 0) / 100, 2)
    taxable     = max(subtotal - disc_amt, 0.0)
    gst_amt     = round(taxable * GST_RATE / 100, 2)
    grand       = round(taxable + gst_amt, 2)
    return subtotal, disc_amt, taxable, gst_amt, grand


# ── Quotation builder ─────────────────────────────────────────────────────────

def build_quotes(client_type, df_app, df_fees,
                 selected_accounting=None, selected_pt_sub=None):
    ct = normalize_str(client_type)
    applicable = (
        df_app.query("ClientType == @ct and Applicable == True")
        .loc[:, ["Service","SubService","ClientType"]].copy()
    )
    if selected_accounting:
        sel    = normalize_str(selected_accounting)
        is_acc = applicable["Service"].eq("ACCOUNTING")
        applicable = pd.concat([
            applicable.loc[~is_acc],
            applicable.loc[is_acc & (applicable["SubService"] == sel)],
        ], ignore_index=True)

    ev_mask   = applicable["Service"].eq(normalize_str(EVENT_SERVICE))
    main_app  = applicable.loc[~ev_mask].copy()
    event_app = applicable.loc[ev_mask].copy()

    is_pt = main_app["Service"].eq(normalize_str(PT_SERVICE))
    if selected_pt_sub is not None:
        sel_pt   = normalize_str(selected_pt_sub)
        main_app = pd.concat([
            main_app.loc[~is_pt],
            main_app.loc[is_pt & (main_app["SubService"] == sel_pt)],
        ], ignore_index=True)

    force_mask = main_app["SubService"].isin(FORCE_EVENT_SUBS)
    if force_mask.any():
        event_app = pd.concat([event_app, main_app.loc[force_mask]], ignore_index=True)
        main_app  = main_app.loc[~force_mask].copy()

    def _fmt(df_in):
        if df_in.empty:
            return pd.DataFrame(columns=["Service","Details","Annual Fees (Rs.)"])
        q = df_in.merge(df_fees, on=["Service","SubService","ClientType"],
                        how="left", validate="1:1")
        q["FeeINR"]     = pd.to_numeric(q["FeeINR"], errors="coerce").fillna(0.0)
        sp              = q["Service"].map(title_with_acronyms)
        q["Service"]    = [svc_display(r, p) for r, p in zip(q["Service"], sp)]
        sbp             = q["SubService"].map(title_with_acronyms)
        q["SubService"] = [sub_display(r, p) for r, p in zip(q["SubService"], sbp)]
        q.sort_values(["Service","SubService"], inplace=True)
        return (q.drop(columns=["ClientType"], errors="ignore")
                  .rename(columns={"SubService":"Details","FeeINR":"Annual Fees (Rs.)"})
                  .loc[:,["Service","Details","Annual Fees (Rs.)"]])

    return _fmt(main_app), _fmt(event_app)


def prep_editor_df(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy().reset_index(drop=True)
    out["Annual Fees (Rs.)"] = out["Annual Fees (Rs.)"].apply(
        lambda x: money_inr(float(x)) if str(x).strip() not in ("","nan") else "")
    out["Include"]     = False   # all deselected by default
    out["MoveToEvent"] = False
    out["Order"]       = range(1, len(out) + 1)
    return out


# ── PDF ───────────────────────────────────────────────────────────────────────

def _grouped_rows(df):
    rows = [["Service","Details","Annual Fees<br/>(Rs.)"]]
    for svc, grp in df.groupby("Service", sort=True):
        g = grp.sort_values("Details"); first = True
        for _, r in g.iterrows():
            rows.append([svc if first else "", r["Details"],
                         money_inr(parse_inr(r["Annual Fees (Rs.)"]))])
            first = False
    return rows


def _event_rows(df):
    rows = [["Details","Fees<br/>(Rs.)"]]
    for _, r in df.iterrows():
        raw    = str(r.get("Annual Fees (Rs.)","")).strip()
        amt    = "" if raw in ("","nan") else money_inr(parse_inr(raw))
        detail = str(r.get("Details","")).strip() or str(r.get("Service","")).strip()
        rows.append([detail, amt])
    return rows


def make_pdf(client_name, client_type, quote_no, df_quote, df_event,
             subtotal, discount_pct, discount_amt, gst_amt, grand,
             letterhead=False, addr="", email="", phone="",
             proposal_start="", discount_reason=""):
    import os
    from reportlab.platypus import Image, HRFlowable
    from reportlab.lib.utils import ImageReader
    from reportlab.lib.enums import TA_RIGHT, TA_LEFT

    buf    = io.BytesIO()
    BB     = colors.HexColor(BRAND_BLUE_HEX)
    LIGHT  = colors.HexColor(BRAND_LIGHT)
    GREY   = colors.HexColor("#F8F9FA")
    DKGREY = colors.HexColor("#4A4A4A")
    BORDER = colors.HexColor("#DEE2E6")

    styles = getSampleStyleSheet()
    hc = ParagraphStyle("HC", parent=styles["Normal"],
                         alignment=TA_CENTER, fontName="Helvetica-Bold", fontSize=9)
    hn = ParagraphStyle("HN", parent=styles["Normal"],
                         fontName="Helvetica-Bold", fontSize=9)
    normal9 = ParagraphStyle("N9", parent=styles["Normal"], fontSize=9)
    normal8 = ParagraphStyle("N8", parent=styles["Normal"], fontSize=8,
                              textColor=DKGREY)
    bold9   = ParagraphStyle("B9", parent=styles["Normal"],
                              fontName="Helvetica-Bold", fontSize=9)
    right9  = ParagraphStyle("R9", parent=styles["Normal"],
                              alignment=TA_RIGHT, fontSize=9)
    bold10  = ParagraphStyle("B10", parent=styles["Normal"],
                              fontName="Helvetica-Bold", fontSize=10)
    title_s = ParagraphStyle("TT", parent=styles["Normal"],
                              fontName="Helvetica-Bold", fontSize=15,
                              textColor=colors.white, alignment=TA_CENTER,
                              spaceAfter=2)
    sub_s   = ParagraphStyle("SS", parent=styles["Normal"],
                              fontSize=9, textColor=colors.HexColor("#CCE0FF"),
                              alignment=TA_CENTER)

    def on_page(canv, doc_):
        pw, ph = A4
        canv.saveState()
        # White background
        canv.setFillColor(colors.white)
        canv.rect(0, 0, pw, ph, stroke=0, fill=1)
        # Top branded header band
        canv.setFillColor(BB)
        canv.rect(0, ph - 28*mm, pw, 28*mm, stroke=0, fill=1)
        # Subtle bottom band
        canv.setFillColor(colors.HexColor("#F0F4F8"))
        canv.rect(0, 0, pw, 22*mm, stroke=0, fill=1)
        # Left accent bar
        canv.setFillColor(colors.HexColor("#1A6BBF"))
        canv.rect(0, 22*mm, 3*mm, ph-50*mm, stroke=0, fill=1)
        if letterhead:
            try:
                for n in ("logo.png","logo.jpg","logo.jpeg"):
                    if os.path.exists(n):
                        ir = ImageReader(n); ow,oh = ir.getSize()
                        tw = pw - 80*mm; r = tw/ow; th = oh*r
                        if hasattr(canv,"setFillAlpha"): canv.setFillAlpha(0.06)
                        canv.drawImage(n,(pw-tw)/2,(ph-th)/2,
                                       width=tw,height=th,
                                       preserveAspectRatio=True,mask="auto")
                        if hasattr(canv,"setFillAlpha"): canv.setFillAlpha(1)
                        break
            except Exception: pass
        canv.restoreState()

    def on_page_end(canv, doc_):
        canv.saveState()
        # Footer text
        canv.setFont("Helvetica", 7.5)
        canv.setFillColor(DKGREY)
        canv.drawCentredString(A4[0]/2, 13*mm,
            "Office No. 5, Ground Floor, Adeshwar Arcade, Andheri-Kurla Road, "
            "Andheri East, Mumbai - 400093")
        canv.drawCentredString(A4[0]/2, 9*mm,
            "Email: info@vpurohit.com  |  Contact: +91-8369508539")
        canv.setFont("Helvetica", 7)
        canv.setFillColor(colors.HexColor("#999999"))
        canv.drawRightString(A4[0]-18*mm, 5*mm, f"Page {canv.getPageNumber()}")
        canv.restoreState()

    doc   = BaseDocTemplate(buf, pagesize=A4,
                            leftMargin=14*mm, rightMargin=14*mm,
                            topMargin=32*mm, bottomMargin=26*mm)
    frame = Frame(doc.leftMargin, doc.bottomMargin,
                  doc.width, doc.height, id="normal")
    doc.addPageTemplates(PageTemplate(id="main", frames=[frame],
                                      onPage=on_page, onPageEnd=on_page_end))
    story = []

    # ── Header content (sits inside top band via negative spacing trick) ──
    # Since onPage draws the band, we just add the text at top of story
    story.append(Spacer(1, 2))

    # Firm name in header — white text over blue band drawn by onPage
    # We use a table to position logo left + firm name right
    hdr_left = []
    for nm in ("logo.png","logo.jpg","logo.jpeg"):
        if os.path.exists(nm) and not letterhead:
            try:
                ir = ImageReader(nm); ow,oh = ir.getSize()
                r  = min(18*mm/ow, 18*mm/oh)
                hdr_left = [Image(nm, width=ow*r, height=oh*r)]
            except Exception:
                pass
            break

    hdr_right = [
        Paragraph("V. Purohit &amp; Associates", title_s),
        Paragraph("Chartered Accountants", sub_s),
    ]

    hdr_tbl = Table(
        [[hdr_left or [Paragraph("", normal9)], hdr_right]],
        colWidths=[30*mm, doc.width - 30*mm]
    )
    hdr_tbl.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING",  (0,0), (-1,-1), 0),
        ("RIGHTPADDING", (0,0), (-1,-1), 0),
        ("TOPPADDING",   (0,0), (-1,-1), 0),
        ("BOTTOMPADDING",(0,0), (-1,-1), 0),
    ]))
    story.append(hdr_tbl)
    story.append(Spacer(1, 8))

    # ── Proposal title bar ──
    title_bar = Table(
        [[Paragraph("ANNUAL FEES PROPOSAL", ParagraphStyle(
            "PT", parent=styles["Normal"],
            fontName="Helvetica-Bold", fontSize=11,
            textColor=BB, alignment=TA_CENTER))]],
        colWidths=[doc.width]
    )
    title_bar.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), LIGHT),
        ("TOPPADDING",    (0,0),(-1,-1), 7),
        ("BOTTOMPADDING", (0,0),(-1,-1), 7),
        ("BOX", (0,0),(-1,-1), 0.8, BB),
    ]))
    story.append(title_bar)
    story.append(Spacer(1, 8))

    # ── Meta info — two column card ──
    lc, rc = [], []
    lc.append(Paragraph(f"<b>Client Name</b>", normal8))
    lc.append(Paragraph(client_name, bold10))
    lc.append(Spacer(1, 4))
    lc.append(Paragraph(f"<b>Entity Type:</b> {client_type}", normal9))
    if addr.strip():
        lc.append(Paragraph(
            "<b>Address:</b> " + ", ".join(l.strip() for l in addr.splitlines() if l.strip()),
            normal9))
    if email.strip():
        lc.append(Paragraph(f"<b>Email:</b> {email.strip()}", normal9))
    if phone.strip():
        lc.append(Paragraph(f"<b>Phone:</b> {phone.strip()}", normal9))

    rc.append(Paragraph("<b>Quotation Details</b>", normal8))
    rc.append(Spacer(1, 2))
    rc.append(Paragraph(f"<b>Ref. No.:</b>  {quote_no}", bold9))
    rc.append(Paragraph(f"<b>Date:</b>  {datetime.now().strftime('%d %B %Y')}", normal9))
    if proposal_start.strip():
        rc.append(Paragraph(f"<b>Proposed Start:</b>  {proposal_start.strip()}", normal9))
    rc.append(Spacer(1, 4))
    rc.append(Paragraph(
        f"<b>Total (incl. GST):</b>",  normal8))
    rc.append(Paragraph(
        f"Rs. {money_inr(grand)}",
        ParagraphStyle("GT", parent=styles["Normal"],
                       fontName="Helvetica-Bold", fontSize=13, textColor=BB)))

    n = max(len(lc), len(rc))
    lc += [Spacer(1,2)] * (n - len(lc))
    rc += [Spacer(1,2)] * (n - len(rc))

    meta_tbl = Table([[lc[i], rc[i]] for i in range(n)],
                      colWidths=[doc.width*0.58, doc.width*0.42])
    meta_tbl.setStyle(TableStyle([
        ("VALIGN",        (0,0),(-1,-1),"TOP"),
        ("LEFTPADDING",   (0,0),(-1,-1), 8),
        ("RIGHTPADDING",  (0,0),(-1,-1), 8),
        ("TOPPADDING",    (0,0),(-1,-1), 6),
        ("BOTTOMPADDING", (0,0),(-1,-1), 6),
        ("BACKGROUND",    (0,0),(0,-1), colors.HexColor("#F8FAFD")),
        ("BACKGROUND",    (1,0),(1,-1), LIGHT),
        ("BOX",           (0,0),(-1,-1), 0.8, BORDER),
        ("LINEAFTER",     (0,0),(0,-1), 0.6, BORDER),
        ("ROUNDEDCORNERS",[3]),
    ]))
    story.append(meta_tbl)
    story.append(Spacer(1, 10))

    # ── Section heading ──
    story.append(Paragraph(
        "SCOPE OF SERVICES &amp; FEES",
        ParagraphStyle("SH", parent=styles["Normal"],
                       fontName="Helvetica-Bold", fontSize=8,
                       textColor=DKGREY, spaceAfter=4)))

    # ── Fees table — use row order as-is (already sorted by Order) ──
    trows = _grouped_rows(df_quote)
    trows[0] = [
        Paragraph("<b>Service</b>", hc),
        Paragraph("<b>Description / Details</b>", hc),
        Paragraph("<b>Annual Fees</b><br/><b>(Rs.)</b>", hc),
    ]
    cw = [55*mm, 85*mm, 33*mm]
    ft = Table(trows, colWidths=cw, repeatRows=1)
    ft.setStyle(TableStyle([
        # Header row
        ("BACKGROUND",    (0,0),(-1,0), BB),
        ("TEXTCOLOR",     (0,0),(-1,0), colors.white),
        ("FONTNAME",      (0,0),(-1,0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),(-1,0), 9),
        ("TOPPADDING",    (0,0),(-1,0), 7),
        ("BOTTOMPADDING", (0,0),(-1,0), 7),
        ("ALIGN",         (0,0),(-1,0), "CENTER"),
        # Data rows
        ("FONTSIZE",      (0,1),(-1,-1), 9),
        ("TOPPADDING",    (0,1),(-1,-1), 5),
        ("BOTTOMPADDING", (0,1),(-1,-1), 5),
        ("LEFTPADDING",   (0,1),(-1,-1), 6),
        ("ALIGN",         (2,1),(2,-1), "RIGHT"),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.white, colors.HexColor("#F4F7FB")]),
        ("INNERGRID",     (0,1),(-1,-1), 0.3, BORDER),
        ("LINEBELOW",     (0,0),(-1,0), 0, colors.white),
        ("BOX",           (0,0),(-1,-1), 0.8, BORDER),
        # Left accent on service column
        ("LINEAFTER",     (0,0),(0,-1), 0.5, BORDER),
    ]))
    story.append(ft)
    story.append(Spacer(1, 6))

    # ── Totals ──
    disc_lbl = f"Discount ({discount_pct:.0f}%)" + (
        f" — {discount_reason}" if discount_reason and discount_reason.strip() else "")
    tot_data = [
        ["", "", ""],   # spacer row
        ["", Paragraph("Subtotal", normal9), Paragraph(money_inr(subtotal), right9)],
        ["", Paragraph(disc_lbl,  normal9),  Paragraph(f"({money_inr(discount_amt)})", right9)]
            if discount_amt > 0 else
        ["", Paragraph("Discount", normal9), Paragraph("Nil", right9)],
        ["", Paragraph("Taxable Amount", bold9),
             Paragraph(money_inr(subtotal - discount_amt),
                       ParagraphStyle("RB9",parent=styles["Normal"],
                                      alignment=TA_RIGHT,fontName="Helvetica-Bold",fontSize=9))],
        ["", Paragraph(f"GST @ {GST_RATE}%", normal9), Paragraph(money_inr(gst_amt), right9)],
        ["", Paragraph("GRAND TOTAL", ParagraphStyle("GT2",parent=styles["Normal"],
                        fontName="Helvetica-Bold",fontSize=10,textColor=BB)),
             Paragraph(f"Rs. {money_inr(grand)}",
                       ParagraphStyle("GTR",parent=styles["Normal"],
                                      alignment=TA_RIGHT,fontName="Helvetica-Bold",
                                      fontSize=10,textColor=BB))],
    ]
    t2 = Table(tot_data, colWidths=cw)
    t2.setStyle(TableStyle([
        ("TOPPADDING",    (0,0),(-1,-1), 3),
        ("BOTTOMPADDING", (0,0),(-1,-1), 3),
        ("LEFTPADDING",   (1,0),(1,-1), 6),
        ("ALIGN",         (2,0),(2,-1), "RIGHT"),
        ("LINEABOVE",     (1,2),(2,2), 0.4, BORDER),
        ("LINEABOVE",     (1,4),(2,4), 0.4, BORDER),
        # Grand total row background
        ("BACKGROUND",    (0,-1),(-1,-1), LIGHT),
        ("TOPPADDING",    (0,-1),(-1,-1), 6),
        ("BOTTOMPADDING", (0,-1),(-1,-1), 6),
        ("BOX",           (0,-1),(-1,-1), 0.8, BB),
    ]))
    story.append(t2)
    story.append(Spacer(1, 10))

    # ── Notes ──
    notes_data = [
        [Paragraph("<b>Notes</b>", bold9)],
        [Paragraph(
            "1. Our scope of engagement is strictly limited to the services enumerated above; "
            "any additional services shall be subject to a separate arrangement.", normal8)],
        [Paragraph(
            f"2. This quotation is valid for a period of <b>15 days</b> from the date of "
            f"generation, i.e., up to <b>{validity_date()}</b>.", normal8)],
        [Paragraph(
            "3. Kindly sign and return this proposal as confirmation of your acceptance of "
            "the terms and fees set out herein.", normal8)],
        [Paragraph(
            "4. Please refer to the <b>'Event-Based Charges'</b> schedule appended hereto "
            "for fees applicable to additional or incidental services.", normal8)],
    ]
    notes_tbl = Table(notes_data, colWidths=[doc.width])
    notes_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), colors.HexColor("#FFFBF0")),
        ("LEFTPADDING",   (0,0),(-1,-1), 8),
        ("RIGHTPADDING",  (0,0),(-1,-1), 8),
        ("TOPPADDING",    (0,0),(-1,-1), 3),
        ("BOTTOMPADDING", (0,0),(-1,-1), 3),
        ("LINEBEFORE",    (0,0),(0,-1), 3, colors.HexColor("#FFC107")),
        ("BOX",           (0,0),(-1,-1), 0.5, BORDER),
    ]))
    story.append(notes_tbl)
    story.append(Spacer(1, 14))

    # ── Signature block ──
    sig_l = Table([
        [Paragraph("For <b>V. Purohit &amp; Associates</b>", bold9)],
        [Paragraph("Chartered Accountants", normal8)],
        [Spacer(1, 18)],
        [Paragraph("_" * 36, normal9)],
        [Paragraph("<b>Authorised Signatory</b>", bold9)],
        [Paragraph(datetime.now().strftime("%d %B %Y"), normal8)],
    ], colWidths=[85*mm])
    sig_l.setStyle(TableStyle([
        ("LEFTPADDING",   (0,0),(-1,-1), 0),
        ("TOPPADDING",    (0,0),(-1,-1), 2),
        ("BOTTOMPADDING", (0,0),(-1,-1), 2),
    ]))

    sig_r = Table([
        [Paragraph(f"Accepted by:", normal8)],
        [Paragraph(f"<b>{client_name}</b>", bold9)],
        [Spacer(1, 18)],
        [Paragraph("_" * 36, normal9)],
        [Paragraph("<b>Signature &amp; Stamp</b>", bold9)],
        [Paragraph("Date: _______________", normal8)],
    ], colWidths=[85*mm])
    sig_r.setStyle(TableStyle([
        ("LEFTPADDING",   (0,0),(-1,-1), 0),
        ("TOPPADDING",    (0,0),(-1,-1), 2),
        ("BOTTOMPADDING", (0,0),(-1,-1), 2),
    ]))

    sig_outer = Table([[sig_l, sig_r]], colWidths=[doc.width/2, doc.width/2])
    sig_outer.setStyle(TableStyle([
        ("VALIGN",        (0,0),(-1,-1), "TOP"),
        ("LEFTPADDING",   (0,0),(-1,-1), 10),
        ("TOPPADDING",    (0,0),(-1,-1), 10),
        ("BOTTOMPADDING", (0,0),(-1,-1), 10),
        ("BACKGROUND",    (0,0),(-1,-1), colors.HexColor("#F8FAFD")),
        ("BOX",           (0,0),(-1,-1), 0.5, BORDER),
        ("LINEAFTER",     (0,0),(0,-1), 0.5, BORDER),
    ]))
    story.append(sig_outer)

    # ── Event-based page ──
    if not df_event.empty:
        story.append(PageBreak())
        story.append(Spacer(1, 6))
        story.append(Paragraph(
            "EVENT-BASED CHARGES",
            ParagraphStyle("SH2", parent=styles["Normal"],
                           fontName="Helvetica-Bold", fontSize=8,
                           textColor=DKGREY, spaceAfter=4)))
        story.append(Paragraph(
            "The following charges are applicable as and when the respective events occur "
            "and are not included in the annual fees quoted above.",
            normal8))
        story.append(Spacer(1, 6))

        erows = _event_rows(df_event)
        erows[0] = [
            Paragraph("<b>Service / Description</b>", hc),
            Paragraph("<b>Applicable Fees (Rs.)</b>", hc),
        ]
        ev = Table(erows, colWidths=[140*mm, 33*mm], repeatRows=1)
        ev.setStyle(TableStyle([
            ("BACKGROUND",    (0,0),(-1,0), BB),
            ("TEXTCOLOR",     (0,0),(-1,0), colors.white),
            ("FONTNAME",      (0,0),(-1,0), "Helvetica-Bold"),
            ("FONTSIZE",      (0,0),(-1,0), 9),
            ("TOPPADDING",    (0,0),(-1,0), 7),
            ("BOTTOMPADDING", (0,0),(-1,0), 7),
            ("ALIGN",         (0,0),(-1,0), "CENTER"),
            ("FONTSIZE",      (0,1),(-1,-1), 9),
            ("TOPPADDING",    (0,1),(-1,-1), 5),
            ("BOTTOMPADDING", (0,1),(-1,-1), 5),
            ("LEFTPADDING",   (0,1),(-1,-1), 6),
            ("ALIGN",         (1,1),(1,-1), "RIGHT"),
            ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
            ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.white, colors.HexColor("#F4F7FB")]),
            ("INNERGRID",     (0,1),(-1,-1), 0.3, BORDER),
            ("BOX",           (0,0),(-1,-1), 0.8, BORDER),
        ]))
        story.append(ev)

    doc.build(story)
    return buf.getvalue()


def export_excel(df_main, df_event, client_name, client_type, quote_no,
                 subtotal, discount_pct, discount_amt, gst_amt, grand,
                 discount_reason=""):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb    = Workbook()
    thin  = Side(style="thin", color="000000")
    ba    = Border(left=thin, right=thin, top=thin, bottom=thin)
    hfill = PatternFill("solid", fgColor=BRAND_BLUE_HEX.replace("#",""))
    hfont = Font(color="FFFFFF", bold=True)
    right = Alignment(horizontal="right")
    ctr   = Alignment(horizontal="center")

    ws = wb.active; ws.title = "Annual Fees"
    ws.append(["Service","Details","Annual Fees (Rs.)"])
    for c in range(1,4):
        cell = ws.cell(1,c)
        cell.fill,cell.font,cell.alignment,cell.border = hfill,hfont,ctr,ba

    for svc, grp in df_main.groupby("Service", sort=True):
        g = grp.sort_values("Details"); first = True
        for _, r in g.iterrows():
            ws.append([svc if first else "",
                        r.get("Details",""),
                        int(round(parse_inr(r.get("Annual Fees (Rs.)","0"))))])
            first = False

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[2].number_format = '#,##,##0'
        for cell in row: cell.border = ba

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 18

    sr = ws.max_row + 2
    ws.cell(sr,2,"Totals").font = Font(bold=True)
    disc_lbl = f"Discount ({int(discount_pct)}%)" + (f" — {discount_reason}" if discount_reason else "")
    for i,(lbl,amt) in enumerate([
        ("Subtotal",subtotal),(disc_lbl,-discount_amt),
        ("Taxable Amount",subtotal-discount_amt),
        ("GST (18%)",gst_amt),("Grand Total",grand),
    ], start=sr+1):
        ws.cell(i,2,lbl)
        c = ws.cell(i,3,int(round(amt)))
        c.number_format = '#,##,##0'; c.alignment = right

    if not df_event.empty:
        ws2 = wb.create_sheet("Event-based charges")
        ws2.append(["Details","Fees (Rs.)"])
        for h in (ws2["A1"],ws2["B1"]):
            h.fill,h.font,h.alignment,h.border = hfill,hfont,ctr,ba
        for _,r in df_event.iterrows():
            detail = str(r.get("Details","")).strip() or str(r.get("Service","")).strip()
            raw    = str(r.get("Annual Fees (Rs.)","")).strip()
            ws2.append([detail, int(round(parse_inr(raw))) if raw else None])
        for row in ws2.iter_rows(min_row=2,max_row=ws2.max_row):
            row[1].number_format = '#,##,##0'
            for cell in row: cell.border = ba
        ws2.column_dimensions["A"].width = 60
        ws2.column_dimensions["B"].width = 18

    ws3 = wb.create_sheet("Cover")
    for r in [["Client Name",client_name],["Entity Type",client_type],
               ["Quotation No.",quote_no],
               ["Date",datetime.now().strftime("%d-%b-%Y")],
               ["Valid Until",validity_date()],["FY",get_fy(datetime.now())]]:
        ws3.append(r)
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 50

    bio = io.BytesIO(); wb.save(bio); return bio.getvalue()


# ── Session state ─────────────────────────────────────────────────────────────

def ss(k, v):
    if k not in st.session_state: st.session_state[k] = v

ss("editor_active",  False)
ss("quote_df",       pd.DataFrame())
ss("event_df",       pd.DataFrame())
ss("client_name",    ""); ss("client_type",  ""); ss("client_id","")
ss("client_addr",    ""); ss("client_email", ""); ss("client_phone","")
ss("proposal_start", ""); ss("quote_no",     "")
ss("discount_pct",   0);  ss("discount_reason",""); ss("letterhead",False)
ss("quote_saved",    False)


# ── Page config & CSS ─────────────────────────────────────────────────────────

st.set_page_config(page_title=APP_TITLE, page_icon="📄", layout="wide")

st.markdown(f"""<style>
:root{{color-scheme:only light;}}
html,body,.stApp,.block-container{{background:#FFFFFF !important;color:#111 !important;}}
h1,h2,h3,h4,h5,h6{{color:#0E0E0E !important;}}
.stButton>button,.stDownloadButton>button{{
  background:{BRAND_BLUE_HEX} !important;color:#fff !important;
  border:0;border-radius:6px;font-weight:600;}}
.stButton>button:hover,.stDownloadButton>button:hover{{background:#0a3a6b !important;}}
h1>a,h2>a,h3>a,h4>a{{display:none !important;}}
.section-divider{{border:none;border-top:2px solid {BRAND_BLUE_HEX};
  margin:16px 0 10px 0;opacity:0.15;}}
.info-box{{background:{BRAND_LIGHT};border-left:4px solid {BRAND_BLUE_HEX};
  padding:10px 14px;border-radius:4px;font-size:0.9rem;margin:8px 0 12px 0;}}
.totals-card{{background:{BRAND_LIGHT};border:1px solid #c9d9f0;
  border-radius:8px;padding:16px 20px;margin:12px 0;}}
.totals-card .grand{{font-size:1.2rem;font-weight:800;color:{BRAND_BLUE_HEX};}}
.brand-header{{background:{BRAND_BLUE_HEX};padding:16px 22px;
  border-radius:8px;margin-bottom:16px;}}
</style>""", unsafe_allow_html=True)


# ── Load Fee Master ───────────────────────────────────────────────────────────

try:
    fee_full     = cached_fee_master()
    df_app       = fee_full[["Service","SubService","ClientType","Applicable"]].copy()
    df_fees      = fee_full[["Service","SubService","ClientType","FeeINR"]].copy()
    CLIENT_TYPES = sorted(df_app["ClientType"].dropna().unique().tolist())
except Exception as e:
    st.error(f"⚠️ Cannot connect to Google Sheets. Check Streamlit Secrets. Error: {e}")
    st.stop()


# ══════════════════════════════════════════════════════════════════════════════
# TABS
# ══════════════════════════════════════════════════════════════════════════════

t1, t2, t3, t4, t5 = st.tabs([
    "📄 Generate Quotation",
    "👥 Client Master",
    "📁 Saved Quotations",
    "📊 Reports",
    "⚙️ Fee Master",
])


# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — GENERATE QUOTATION
# ══════════════════════════════════════════════════════════════════════════════

with t1:
    st.markdown(f"""
    <div class="brand-header">
      <h2 style="color:white;margin:0;font-size:1.3rem;">📄 Generate Quotation</h2>
      <p style="color:#cce0ff;margin:4px 0 0 0;font-size:0.85rem;">
        V. Purohit &amp; Associates — Chartered Accountants
      </p>
    </div>""", unsafe_allow_html=True)

    # Step indicator
    active = 2 if st.session_state["editor_active"] else 1
    sc = st.columns(3)
    for i,(col,lbl) in enumerate(zip(sc,
            ["① Client Details","② Review & Edit Fees","③ Download"]),start=1):
        bg = BRAND_BLUE_HEX if i==active else ("#28a745" if i<active else "#dee2e6")
        tc = "white" if i<=active else "#666"
        col.markdown(
            f'<div style="text-align:center;background:{bg};color:{tc};'
            f'padding:7px 4px;border-radius:6px;font-size:0.82rem;font-weight:600;">{lbl}</div>',
            unsafe_allow_html=True)

    st.markdown("<hr class='section-divider'>", unsafe_allow_html=True)
    left_col, right_col = st.columns([3,2])

    with left_col:
        st.markdown("#### 👤 Client Details")

        clients_df = load_clients()
        cl_names   = (["— Select existing client or type new —"] +
                      sorted(clients_df["ClientName"].dropna().tolist())
                      if not clients_df.empty
                      else ["— No clients yet. Add in Client Master tab —"])

        sel_client = st.selectbox("Search / Select Existing Client", cl_names,
                                   help="Auto-fills all fields below")

        if (sel_client not in ("— Select existing client or type new —",
                                "— No clients yet. Add in Client Master tab —")
                and not clients_df.empty):
            row = clients_df[clients_df["ClientName"] == sel_client].iloc[0]
            for k,v in [("client_name","ClientName"),("client_type","ClientType"),
                        ("client_addr","Address"),("client_email","Email"),
                        ("client_phone","Phone"),("client_id","ClientID")]:
                st.session_state[k] = str(row.get(v,""))

        with st.form("quote_form", clear_on_submit=False):
            ca, cb = st.columns([3,2])
            with ca:
                client_name = st.text_input("Client Name *",
                    st.session_state.get("client_name",""),
                    placeholder="Full legal name")
            with cb:
                ctl  = CLIENT_TYPES
                ctsv = st.session_state.get("client_type","")
                ctix = ctl.index(ctsv) if ctsv in ctl else 0
                client_type = st.selectbox("Client Type *", ctl, index=ctix)

            addr = st.text_area("Address", st.session_state.get("client_addr",""),
                                 placeholder="Street, Area\nCity, State, PIN", height=80)
            cc, cd = st.columns(2)
            with cc:
                email = st.text_input("Email", st.session_state.get("client_email",""),
                                       placeholder="client@example.com")
            with cd:
                phone = st.text_input("Phone", st.session_state.get("client_phone",""),
                                       placeholder="+91 98765 43210")

            proposal_start = st.text_input("Proposed Engagement Start",
                st.session_state.get("proposal_start",""),
                placeholder="e.g., 1st April 2025 / FY 2025-26")

            st.markdown("<hr class='section-divider'>", unsafe_allow_html=True)
            st.markdown("#### 📋 Service Selection")

            ct_norm = normalize_str(client_type)
            app_ct  = df_app[(df_app["ClientType"]==ct_norm) & (df_app["Applicable"]==True)]
            ce, cf  = st.columns(2)
            with ce:
                st.markdown("**Accounting Plan**")
                sel_acc = st.radio("Accounting Plan", ACCOUNTING_PLANS,
                    index=3, horizontal=False, label_visibility="collapsed")
            with cf:
                st.markdown("**Profession Tax Returns**")
                pt_opts = sorted([
                    title_with_acronyms(s)
                    for s in app_ct.loc[
                        app_ct["Service"]==normalize_str(PT_SERVICE),"SubService"
                    ].dropna().unique().tolist() if s
                ])
                sel_pt = st.radio("PT Returns",
                    pt_opts if pt_opts else ["(Not applicable)"],
                    index=0, horizontal=False, label_visibility="collapsed")
                if sel_pt == "(Not applicable)": sel_pt = None

            submit = st.form_submit_button("⚡ Generate Quotation Table",
                                            use_container_width=True)

    with right_col:
        st.markdown("#### ⚙️ Proposal Options")
        st.session_state["discount_pct"] = st.number_input(
            "Discount %", 0, 100, int(st.session_state["discount_pct"]), 1)
        if st.session_state["discount_pct"] > 0:
            st.session_state["discount_reason"] = st.text_input(
                "Reason for discount",
                st.session_state.get("discount_reason",""),
                placeholder="e.g., Introductory offer")
        st.session_state["letterhead"] = st.checkbox(
            "Letterhead mode (watermark logo)",
            value=st.session_state["letterhead"])
        st.markdown("<hr class='section-divider'>", unsafe_allow_html=True)
        st.markdown("#### 💡 Quick Guide")
        st.markdown("""
1. Select existing client **or** type new details
2. Choose accounting plan & PT type
3. Click **Generate**
4. Edit fees → Download PDF / Excel
5. Click **💾 Save** to log the quotation
        """)
        if not clients_df.empty:
            st.metric("Clients on Record", len(clients_df))

    # Submit
    if submit:
        errs = []
        if not client_name.strip(): errs.append("Client Name is required.")
        if not validate_email(email): errs.append(f"Invalid email: {email}")
        if not validate_phone(phone): errs.append(f"Invalid phone: {phone}")
        if errs:
            for e in errs: st.error(f"⚠️ {e}")
        else:
            qno = generate_quote_no()
            st.session_state.update({
                "quote_no":qno,"proposal_start":proposal_start,
                "client_name":client_name,"client_type":client_type,
                "client_addr":addr,"client_email":email,
                "client_phone":phone,"quote_saved":False,
            })
            main_df, event_df = build_quotes(
                client_type, df_app, df_fees,
                selected_accounting=sel_acc, selected_pt_sub=sel_pt)

            cons = pd.DataFrame([{"Service":"Consulting Charges",
                                   "Details":"Consulting Charges",
                                   "Annual Fees (Rs.)":""}])
            if event_df.empty:
                event_df = cons
            elif not event_df["Details"].str.strip().str.casefold().eq("consulting charges").any():
                event_df = pd.concat([event_df, cons], ignore_index=True)

            if main_df.empty and event_df.empty:
                st.warning("No applicable services found.")
            else:
                ev = event_df.copy()
                ev["Annual Fees (Rs.)"] = ev["Annual Fees (Rs.)"].apply(
                    lambda v: "" if str(v).strip() in ("","nan")
                    else money_inr(parse_inr(str(v).replace(",",""))))
                ev["MoveToMain"] = False
                st.session_state["quote_df"]      = prep_editor_df(main_df)
                st.session_state["event_df"]      = ev
                st.session_state["editor_active"] = True

    # Editor
    if st.session_state["editor_active"] and (
        not st.session_state["quote_df"].empty or
        not st.session_state["event_df"].empty
    ):
        st.markdown("<hr class='section-divider'>", unsafe_allow_html=True)
        h1, h2 = st.columns([5,1])
        h1.markdown(f"#### 📝 {st.session_state['client_name']} — "
                    f"{st.session_state['client_type'].title()}")
        if h2.button("↩ Start Over"):
            st.session_state.update({
                "editor_active":False,
                "quote_df":pd.DataFrame(),
                "event_df":pd.DataFrame(),
            })
            st.rerun()

        st.markdown('<div class="info-box">'
                    '✅ <b>Check</b> the services you want to include. '
                    'Edit the <b>Order</b> number to rearrange rows in the PDF. '
                    'Use <b>→ Event</b> to move a service to the event-based section.</div>',
                    unsafe_allow_html=True)

        # Main editor
        if not st.session_state["quote_df"].empty:
            with st.form("edit_main"):
                edited = st.data_editor(
                    st.session_state["quote_df"],
                    use_container_width=True,
                    disabled=["Service","Details"],
                    column_order=["Order","Include","MoveToEvent","Service","Details","Annual Fees (Rs.)"],
                    column_config={
                        "Order": st.column_config.NumberColumn(
                            "Order", help="Edit number to rearrange rows in PDF output.",
                            min_value=1, step=1, width="small"),
                        "Include":     st.column_config.CheckboxColumn(
                            "Include", help="Check to include in proposal.", width="small"),
                        "MoveToEvent": st.column_config.CheckboxColumn(
                            "→ Event", help="Move to event-based section.", width="small"),
                        "Service": st.column_config.TextColumn("Service", width="medium"),
                        "Details": st.column_config.TextColumn("Details", width="medium"),
                        "Annual Fees (Rs.)": st.column_config.TextColumn(
                            "Annual Fees (Rs.)", validate=r"^\s*[\d,]*\s*$", width="small"),
                    },
                    num_rows="fixed", key="qeditor", hide_index=True, height=420)
                b1, b2 = st.columns(2)
                do_apply = b1.form_submit_button("✅ Apply Edits", use_container_width=True)
                do_move  = b2.form_submit_button("✅ Apply & Move to Event", use_container_width=True)
                if do_apply or do_move:
                    edited["Annual Fees (Rs.)"] = edited["Annual Fees (Rs.)"].apply(
                        lambda x: money_inr(parse_inr(x)))
                    # Sort by Order column before saving
                    edited = edited.sort_values("Order").reset_index(drop=True)
                    st.session_state["quote_df"] = edited.copy()
                    if do_move:
                        mv = edited[edited["MoveToEvent"]].copy()
                        if not mv.empty:
                            mv = mv[["Service","Details","Annual Fees (Rs.)"]].copy()
                            mv["MoveToMain"] = False
                            st.session_state["event_df"] = pd.concat(
                                [st.session_state["event_df"], mv], ignore_index=True)
                            kept = edited[~edited["MoveToEvent"]].copy()
                            kept["MoveToEvent"] = False
                            st.session_state["quote_df"] = kept
                            st.success(f"Moved {len(mv)} row(s) to Event-based.")

        qdf      = st.session_state["quote_df"]
        filtered = qdf[qdf["Include"]==True].sort_values("Order").drop(
            columns=["Include","MoveToEvent","Order"], errors="ignore").copy()

        # Event editor
        ev_df = st.session_state["event_df"].copy()
        if not ev_df.empty:
            st.markdown("<hr class='section-divider'>", unsafe_allow_html=True)
            st.markdown("#### 📌 Event-Based Charges")
            with st.form("event_form"):
                ev_ed = st.data_editor(
                    ev_df, use_container_width=True,
                    disabled=["Service","Details"],
                    column_order=["MoveToMain","Service","Details","Annual Fees (Rs.)"],
                    column_config={
                        "MoveToMain": st.column_config.CheckboxColumn("→ Main"),
                        "Annual Fees (Rs.)": st.column_config.TextColumn(
                            "Fees (Rs.)", validate=r"^\s*[\d,]*\s*$"),
                    },
                    num_rows="fixed", key="eeditor", hide_index=True, height=280)
                if st.form_submit_button("✅ Apply Event Edits", use_container_width=True):
                    def _ef(v):
                        s = str(v).strip()
                        return "" if s in ("","nan") else money_inr(parse_inr(s.replace(",","")))
                    ev_ed["Annual Fees (Rs.)"] = ev_ed["Annual Fees (Rs.)"].apply(_ef)
                    to_main = ev_ed[ev_ed["MoveToMain"]].copy()
                    keep_ev = ev_ed[~ev_ed["MoveToMain"]].copy()
                    if not to_main.empty:
                        add = to_main[["Service","Details","Annual Fees (Rs.)"]].copy()
                        add["Include"] = True; add["MoveToEvent"] = False
                        st.session_state["quote_df"] = pd.concat(
                            [st.session_state["quote_df"], add], ignore_index=True)
                        st.success(f"Moved {len(add)} row(s) to Main.")
                    keep_ev["MoveToMain"] = False
                    st.session_state["event_df"] = keep_ev

        # Totals
        st.markdown("<hr class='section-divider'>", unsafe_allow_html=True)
        calc = filtered.copy()
        calc["Annual Fees (Rs.)"] = calc["Annual Fees (Rs.)"].apply(parse_inr)
        subtotal, disc_amt, taxable, gst_amt, grand = compute_totals(
            calc, st.session_state["discount_pct"])
        dp = st.session_state["discount_pct"]
        dr = st.session_state.get("discount_reason","")

        st.markdown(f"""
        <div class="totals-card">
          <div style="display:flex;justify-content:space-between;margin-bottom:4px;">
            <span>Subtotal</span><span><b>Rs. {money_inr(subtotal)}</b></span></div>
          <div style="display:flex;justify-content:space-between;margin-bottom:4px;color:#666;">
            <span>Discount ({dp}%){" — "+dr if dr else ""}</span>
            <span>- Rs. {money_inr(disc_amt)}</span></div>
          <div style="display:flex;justify-content:space-between;margin-bottom:4px;">
            <span>Taxable Amount</span><span><b>Rs. {money_inr(taxable)}</b></span></div>
          <div style="display:flex;justify-content:space-between;margin-bottom:8px;color:#666;">
            <span>GST @ 18%</span><span>Rs. {money_inr(gst_amt)}</span></div>
          <hr style="border:none;border-top:1.5px solid {BRAND_BLUE_HEX};opacity:0.3;margin:8px 0;">
          <div style="display:flex;justify-content:space-between;" class="grand">
            <span>Grand Total</span><span>Rs. {money_inr(grand)}</span></div>
          <div style="color:#888;font-size:0.8rem;margin-top:6px;">
            Valid until: <b>{validity_date()}</b> &nbsp;|&nbsp; {get_fy(datetime.now())}
          </div>
        </div>""", unsafe_allow_html=True)

        # Downloads + Save + WhatsApp
        st.markdown("#### ⬇️ Download & Save")
        d1, d2, d3 = st.columns(3)

        with d1:
            try:
                xls = export_excel(
                    filtered, st.session_state["event_df"],
                    st.session_state["client_name"], st.session_state["client_type"],
                    st.session_state["quote_no"],
                    subtotal, dp, disc_amt, gst_amt, grand, dr)
                st.download_button("📊 Download Excel", data=xls,
                    file_name=f"Proposal_{st.session_state['client_name'].replace(' ','_')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
            except Exception:
                st.warning("Excel export unavailable.")

        with d2:
            pdf = make_pdf(
                st.session_state["client_name"], st.session_state["client_type"],
                st.session_state["quote_no"], filtered, st.session_state["event_df"],
                subtotal, dp, disc_amt, gst_amt, grand,
                letterhead=st.session_state["letterhead"],
                addr=st.session_state["client_addr"],
                email=st.session_state["client_email"],
                phone=st.session_state["client_phone"],
                proposal_start=st.session_state["proposal_start"],
                discount_reason=dr)
            st.download_button("📄 Download PDF", data=pdf,
                file_name=f"Proposal_{st.session_state['client_name'].replace(' ','_')}.pdf",
                mime="application/pdf", use_container_width=True)

        with d3:
            if not st.session_state.get("quote_saved", False):
                if st.button("💾 Save Quotation", use_container_width=True):
                    try:
                        cid = st.session_state.get("client_id","") or str(uuid.uuid4())[:8]
                        save_quotation(
                            st.session_state["quote_no"], cid,
                            st.session_state["client_name"],
                            st.session_state["client_type"],
                            filtered, subtotal, dp, dr, gst_amt, grand,
                            st.session_state["proposal_start"])
                        st.session_state["quote_saved"] = True
                        st.success("✅ Saved!")
                        st.cache_data.clear()
                    except Exception as ex:
                        st.error(f"Save failed: {ex}")
            else:
                st.success("✅ Saved to Sheets")

        if grand > 0:
            cn  = st.session_state["client_name"]
            msg = (f"Dear {cn},%0A%0APlease find our annual fees proposal.%0A%0A"
                   f"Grand Total (incl. GST): Rs. {money_inr(grand)}%0A"
                   f"Valid Until: {validity_date()}%0A%0AV. Purohit %26 Associates")
            st.markdown(
                f'<a href="https://wa.me/?text={msg}" target="_blank" '
                f'style="display:inline-block;background:#25D366;color:white;'
                f'padding:8px 18px;border-radius:6px;text-decoration:none;'
                f'font-weight:600;font-size:0.9rem;margin-top:8px;">'
                f'💬 Share via WhatsApp</a>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — CLIENT MASTER
# ══════════════════════════════════════════════════════════════════════════════

with t2:
    st.markdown("### 👥 Client Master")
    st.caption("Add, edit, or delete clients. All data saved live to Google Sheets.")

    cl_df = load_clients()

    with st.expander("➕ Add New Client", expanded=cl_df.empty):
        with st.form("add_client"):
            na, nb = st.columns([3,2])
            with na: nc_name = st.text_input("Client Name *")
            with nb: nc_type = st.selectbox("Client Type *", CLIENT_TYPES)
            nc_addr = st.text_area("Address", height=70,
                                    placeholder="Street, Area\nCity, State, PIN")
            nc, nd  = st.columns(2)
            with nc: nc_email = st.text_input("Email", placeholder="client@example.com")
            with nd: nc_phone = st.text_input("Phone", placeholder="+91 98765 43210")
            if st.form_submit_button("💾 Save Client", use_container_width=True):
                errs = []
                if not nc_name.strip(): errs.append("Name is required.")
                if not validate_email(nc_email): errs.append("Invalid email.")
                if not validate_phone(nc_phone): errs.append("Invalid phone.")
                if errs:
                    for e in errs: st.error(e)
                else:
                    existing = (cl_df["ClientName"].str.strip().str.upper().tolist()
                                if not cl_df.empty else [])
                    if nc_name.strip().upper() in existing:
                        st.warning(f"'{nc_name}' already exists.")
                    else:
                        save_new_client({
                            "ClientID":  str(uuid.uuid4())[:8],
                            "ClientName":nc_name.strip(),
                            "ClientType":nc_type,
                            "Address":   nc_addr.strip(),
                            "Email":     nc_email.strip(),
                            "Phone":     nc_phone.strip(),
                            "CreatedOn": datetime.now().strftime("%d-%b-%Y"),
                        })
                        st.success(f"✅ '{nc_name}' added.")
                        st.cache_data.clear()
                        st.rerun()

    st.markdown("<hr class='section-divider'>", unsafe_allow_html=True)

    if cl_df.empty:
        st.info("No clients yet. Add your first client above.")
    else:
        srch = st.text_input("🔍 Search", placeholder="Name or type...")
        vdf  = cl_df.copy()
        if srch.strip():
            msk = (vdf["ClientName"].str.contains(srch,case=False,na=False) |
                   vdf["ClientType"].str.contains(srch,case=False,na=False))
            vdf = vdf[msk]
        st.markdown(f"**{len(vdf)} client(s)**")

        ed_cl = st.data_editor(vdf, use_container_width=True,
                                disabled=["ClientID","CreatedOn"],
                                hide_index=True, num_rows="dynamic",
                                key="client_editor")

        sc1, sc2 = st.columns(2)
        with sc1:
            if st.button("💾 Save All Changes", use_container_width=True):
                try:
                    if srch.strip():
                        full = cl_df.copy().set_index("ClientID")
                        for _, r in ed_cl.iterrows():
                            if r["ClientID"] in full.index:
                                full.loc[r["ClientID"]] = r
                        final = full.reset_index()
                    else:
                        final = ed_cl
                    overwrite_clients(final)
                    st.success("✅ Client Master updated.")
                    st.cache_data.clear()
                except Exception as ex:
                    st.error(f"Save failed: {ex}")
        with sc2:
            xls_cl = io.BytesIO()
            cl_df.to_excel(xls_cl, index=False)
            st.download_button("📥 Export to Excel", data=xls_cl.getvalue(),
                file_name="ClientMaster.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — SAVED QUOTATIONS
# ══════════════════════════════════════════════════════════════════════════════

with t3:
    st.markdown("### 📁 Saved Quotations")
    st.caption("Browse all past quotations. Load any previous quote to reuse or renew.")

    q_df = load_quotations()

    if q_df.empty:
        st.info("No saved quotations yet. Generate one from the Quotation tab and click 💾 Save.")
    else:
        fc1, fc2, fc3 = st.columns(3)
        with fc1:
            fy_opts = ["All"] + sorted(q_df["FY"].dropna().unique().tolist(), reverse=True)
            r_fy    = st.selectbox("Financial Year", fy_opts, key="h_fy")
        with fc2:
            ct_opts = ["All"] + sorted(q_df["ClientType"].dropna().unique().tolist())
            r_ct    = st.selectbox("Client Type", ct_opts, key="h_ct")
        with fc3:
            cn_opts = ["All"] + sorted(q_df["ClientName"].dropna().unique().tolist())
            r_cn    = st.selectbox("Client", cn_opts, key="h_cn")

        view = q_df.copy()
        if r_fy != "All": view = view[view["FY"]         == r_fy]
        if r_ct != "All": view = view[view["ClientType"] == r_ct]
        if r_cn != "All": view = view[view["ClientName"] == r_cn]

        summary = (view.drop_duplicates("QuoteNo")
                       [["QuoteNo","Date","ClientName","ClientType",
                          "Subtotal","GrandTotal","Status","FY"]]
                       .sort_values("Date", ascending=False).copy())
        summary["Subtotal"]   = summary["Subtotal"].apply(lambda x: f"Rs. {money_inr(x)}")
        summary["GrandTotal"] = summary["GrandTotal"].apply(lambda x: f"Rs. {money_inr(x)}")

        st.markdown(f"**{len(summary)} quotation(s) found**")
        st.dataframe(summary, use_container_width=True, hide_index=True)

        st.markdown("<hr class='section-divider'>", unsafe_allow_html=True)
        st.markdown("#### 🔄 Load Previous Quotation")
        st.caption("Select any past quotation to load it into the Generate tab "
                   "(with last year's fees — edit freely before saving as a new quote).")

        all_qnos = sorted(q_df["QuoteNo"].unique().tolist(), reverse=True)
        sel_qno  = st.selectbox("Quotation No.", ["— select —"] + all_qnos)

        if sel_qno != "— select —":
            prev = q_df[q_df["QuoteNo"] == sel_qno].copy()
            if not prev.empty:
                meta = prev.iloc[0]
                st.markdown(
                    f"**Client:** {meta['ClientName']} &nbsp;|&nbsp; "
                    f"**Type:** {meta['ClientType']} &nbsp;|&nbsp; "
                    f"**Date:** {meta['Date']} &nbsp;|&nbsp; "
                    f"**Grand Total:** Rs. {money_inr(meta['GrandTotal'])}")

                fp = prev[["Service","Details","FeeINR"]].rename(
                    columns={"FeeINR":"Annual Fees (Rs.)"})
                st.dataframe(fp, use_container_width=True, hide_index=True)

                if st.button("📋 Load into Generator", use_container_width=True):
                    rebuild = fp.copy()
                    rebuild["Annual Fees (Rs.)"] = rebuild["Annual Fees (Rs.)"].apply(money_inr)
                    rebuild["Include"]     = False
                    rebuild["MoveToEvent"] = False
                    rebuild["Order"]       = range(1, len(rebuild) + 1)

                    cl_df2 = load_clients()
                    cl_match = (cl_df2[cl_df2["ClientName"].str.strip().str.upper()
                                       == str(meta["ClientName"]).strip().upper()]
                                if not cl_df2.empty else pd.DataFrame())

                    st.session_state.update({
                        "client_name":   meta["ClientName"],
                        "client_type":   meta["ClientType"],
                        "client_id":     cl_match.iloc[0]["ClientID"] if not cl_match.empty else "",
                        "client_addr":   cl_match.iloc[0]["Address"]  if not cl_match.empty else "",
                        "client_email":  cl_match.iloc[0]["Email"]    if not cl_match.empty else "",
                        "client_phone":  cl_match.iloc[0]["Phone"]    if not cl_match.empty else "",
                        "quote_df":      rebuild,
                        "event_df":      pd.DataFrame(columns=["Service","Details",
                                                                "Annual Fees (Rs.)","MoveToMain"]),
                        "editor_active": True,
                        "quote_no":      generate_quote_no(),
                        "quote_saved":   False,
                        "proposal_start":get_fy(datetime.now()),
                    })
                    st.success("✅ Loaded! Switch to the '📄 Generate Quotation' tab to review.")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 4 — REPORTS
# ══════════════════════════════════════════════════════════════════════════════

with t4:
    st.markdown("### 📊 Quotation Reports")

    rpt = load_reports()

    if rpt.empty:
        st.info("No data to report yet. Save some quotations first.")
    else:
        rc1, rc2, rc3 = st.columns(3)
        with rc1:
            fy_r = ["All"] + sorted(rpt["FY"].dropna().unique().tolist(), reverse=True)
            r_fy = st.selectbox("Financial Year", fy_r, key="rpt_fy")
        with rc2:
            ct_r = ["All"] + sorted(rpt["ClientType"].dropna().unique().tolist())
            r_ct = st.selectbox("Client Type", ct_r, key="rpt_ct")
        with rc3:
            st_r = ["All"] + sorted(rpt["Status"].dropna().unique().tolist())
            r_st = st.selectbox("Status", st_r, key="rpt_st")

        rv = rpt.copy()
        if r_fy != "All": rv = rv[rv["FY"]         == r_fy]
        if r_ct != "All": rv = rv[rv["ClientType"] == r_ct]
        if r_st != "All": rv = rv[rv["Status"]     == r_st]

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Quotations",   len(rv))
        m2.metric("Subtotal",     f"Rs. {money_inr(rv['Subtotal'].sum())}")
        m3.metric("GST",          f"Rs. {money_inr(rv['GSTAmt'].sum())}")
        m4.metric("Grand Total",  f"Rs. {money_inr(rv['GrandTotal'].sum())}")

        st.markdown("<hr class='section-divider'>", unsafe_allow_html=True)
        st.markdown("#### By Client Type")
        grp = (rv.groupby("ClientType")
                  .agg(Quotations=("QuoteNo","count"), Total=("GrandTotal","sum"))
                  .reset_index().sort_values("Total", ascending=False))
        grp["Total"] = grp["Total"].apply(lambda x: f"Rs. {money_inr(x)}")
        st.dataframe(grp, use_container_width=True, hide_index=True)

        st.markdown("#### All Quotations")
        disp = rv.copy()
        for col in ["Subtotal","GSTAmt","GrandTotal"]:
            disp[col] = disp[col].apply(lambda x: f"Rs. {money_inr(x)}")
        st.dataframe(disp, use_container_width=True, hide_index=True)

        xls_r = io.BytesIO()
        rv.to_excel(xls_r, index=False)
        st.download_button("📥 Export Report to Excel", data=xls_r.getvalue(),
            file_name=f"QuotationReport_{r_fy.replace(' ','_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 5 — FEE MASTER
# ══════════════════════════════════════════════════════════════════════════════

with t5:
    st.markdown("### ⚙️ Fee Master")
    st.caption("Edit standard fees. Changes apply to all future quotations.")
    st.markdown(
        '<div class="info-box">⚠️ Editing fees here updates the Google Sheet immediately. '
        'Existing saved quotations are unaffected.</div>',
        unsafe_allow_html=True)

    raw_fm = load_fee_master_raw()

    if raw_fm.empty:
        st.warning("Fee Master is empty. Populate the FeeMaster tab in your Google Sheet first.")
    else:
        fm1, fm2 = st.columns(2)
        with fm1:
            f_ct  = st.selectbox("Filter: Client Type",
                ["All"] + sorted(raw_fm["ClientType"].unique().tolist()), key="f_ct")
        with fm2:
            f_svc = st.selectbox("Filter: Service",
                ["All"] + sorted(raw_fm["Service"].unique().tolist()), key="f_svc")

        vfm = raw_fm.copy()
        if f_ct  != "All": vfm = vfm[vfm["ClientType"] == f_ct]
        if f_svc != "All": vfm = vfm[vfm["Service"]    == f_svc]
        st.markdown(f"**{len(vfm)} rows** displayed")

        ed_fm = st.data_editor(
            vfm, use_container_width=True,
            disabled=["Service","SubService","ClientType"],
            column_config={
                "Applicable": st.column_config.CheckboxColumn("Applicable"),
                "FeeINR":     st.column_config.NumberColumn(
                    "Standard Fee (Rs.)", min_value=0, step=500, format="₹%d"),
            },
            hide_index=True, key="fee_editor")

        fs1, fs2 = st.columns(2)
        with fs1:
            if st.button("💾 Save Fee Changes to Google Sheets", use_container_width=True):
                try:
                    full_fm = raw_fm.copy().set_index(
                        ["Service","SubService","ClientType"])
                    for _, r in ed_fm.iterrows():
                        idx = (r["Service"],r["SubService"],r["ClientType"])
                        if idx in full_fm.index:
                            full_fm.loc[idx,"Applicable"] = r["Applicable"]
                            full_fm.loc[idx,"FeeINR"]     = r["FeeINR"]
                    save_fee_master(full_fm.reset_index())
                    st.success("✅ Fee Master saved to Google Sheets.")
                    st.cache_data.clear()
                except Exception as ex:
                    st.error(f"Save failed: {ex}")
        with fs2:
            xls_fm = io.BytesIO()
            raw_fm.to_excel(xls_fm, index=False)
            st.download_button("📥 Download Fee Master (Excel)",
                data=xls_fm.getvalue(), file_name="FeeMaster.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)

        st.markdown("<hr class='section-divider'>", unsafe_allow_html=True)
        st.markdown("#### Standard Fee Summary by Client Type")
        fsumm = (raw_fm[raw_fm["Applicable"]==True]
                 .groupby("ClientType")["FeeINR"].sum()
                 .reset_index()
                 .rename(columns={"FeeINR":"Total (Rs.)"})
                 .sort_values("Total (Rs.)", ascending=False))
        fsumm["Total (Rs.)"] = fsumm["Total (Rs.)"].apply(lambda x: f"Rs. {money_inr(x)}")
        st.dataframe(fsumm, use_container_width=True, hide_index=True)
